import os
from typing import List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from model.Cell import Cell
from model.Layer import Layer
from model.Board import ModelBoard


# Function that reads Excel file Von Neumann's sliding window rule
def read_von_neumann_excel(sheet: Worksheet, start_row: int, start_column: int,
                           cells_states: List[Tuple[str, str]]) -> List[int]:
    zero_color = sheet.cell(start_row + 1, start_column).fill.start_color.value
    zero_color = "#" + zero_color[2:]
    zero_type = 0
    for i in range(len(cells_states)):
        if cells_states[i][1] == zero_color: zero_type = i
    one_color = sheet.cell(start_row, start_column + 1).fill.start_color.value
    one_color = "#" + one_color[2:]
    one_type = 0
    for i in range(len(cells_states)):
        if cells_states[i][1] == one_color: one_type = i
    two_color = sheet.cell(start_row + 1, start_column + 1).fill.start_color.value
    two_color = "#" + two_color[2:]
    two_type = 0
    for i in range(len(cells_states)):
        if cells_states[i][1] == two_color: two_type = i
    three_color = sheet.cell(start_row + 2, start_column + 1).fill.start_color.value
    three_color = "#" + three_color[2:]
    three_type = 0
    for i in range(len(cells_states)):
        if cells_states[i][1] == three_color: three_type = i
    four_color = sheet.cell(start_row + 1, start_column + 2).fill.start_color.value
    four_color = "#" + four_color[2:]
    four_type = 0
    for i in range(len(cells_states)):
        if cells_states[i][1] == four_color: four_type = i
    rule = [zero_type, one_type, two_type, three_type, four_type]
    return rule


# Function that reads Excel file Moore's sliding window rule
def read_moore_excel(sheet: Worksheet, start_row: int, start_column: int,
                     cells_states: List[Tuple[str, str]]) -> List[int]:
    rule = []
    for x in range(3):
        for y in range(3):
            color = sheet.cell(start_row + x, start_column + y).fill.start_color.value
            color = "#" + color[2:]
            typee = 0
            for i in range(len(cells_states)):
                if cells_states[i][1] == color: typee = i
            rule.append(typee)
    return rule


# Function that reads Text file and returns layer object
def read_layer_txt(name: str) -> Layer:
    example = None
    if name == "game of life":
        example = name
        name = "Examples/Game of Life/text/layer1.txt"
    elif name == "Langton ant":
        example = name
        name = "Examples/Langton Ant/text/layer1.txt"
    elif name == "wireworld":
        example = name
        name = "Examples/Wire World/text/layer1.txt"

    with open(name) as reader:

        name = reader.readline().strip()
        size = reader.readline().strip().split()
        height = int(size[1])
        width = int(size[0])
        cells = [[0 for _ in range(width)] for _ in range(height)]
        cells_states = []
        neighbourhood = reader.readline().strip()

        # read grid
        reader.readline()
        for y in range(height):
            line = reader.readline().strip()
            for x in range(width):
                cells[y][x] = int(line[x])

        # read cells states
        reader.readline()
        line = reader.readline()
        cells_states.append(("nothing", "white"))
        while line != "":
            line = line.strip().split()
            cells_states.append((line[0], line[1]))
            line = reader.readline()

        # read cells values
        line = reader.readline()
        values = list()
        while line != "":
            line = line.strip()
            values.append(float(line))
            line = reader.readline()

        # Create layer with all parameters
        cells_class: List = list()
        i = 0
        for y in range(height):
            cells_class.append(list())
            for x in range(width):
                # Check if values have been declared
                if len(values) == 0:
                    cells_class[y].append(Cell(cells[y][x], example=example))
                else:
                    # Check if current cell had value declared
                    if cells[y][x] != 0:
                        cells_class[y].append(Cell(cells[y][x], value=values[i], example=example))
                        i += 1
                    else:
                        cells_class[y].append(Cell(cells[y][x], example=example))

        layer = Layer(name, cells_class, neighbourhood, cells_states)
        return layer


# Function that reads Excel file and returns layer list
def read_layer_excel(name: str) -> List[Layer]:
    example = None

    if name == "forest fire":
        example = name
        name = "Examples/Forest Fire/excel/example.xlsx"
    elif name == "game of life":
        example = name
        name = "Examples/Game of Life/excel/example.xlsx"
    elif name == "Langton ant":
        example = name
        name = "Examples/Langton Ant/excel/example.xlsx"
    elif name == "wireworld":
        example = name
        name = "Examples/Wire World/excel/example.xlsx"

    workbook = load_workbook(filename=name)
    sheets_names = workbook.get_sheet_names()
    layers: List[Layer] = list()

    for sheet_name in sheets_names:
        sheet = workbook.get_sheet_by_name(sheet_name)

        name = sheet_name
        height = int(sheet["B1"].value)
        width = int(sheet["A1"].value)
        neighbourhood = sheet["A2"].value
        cells_states = []
        cells = [[tuple() for _ in range(width)] for _ in range(height)]

        # read cell types
        cells_states.append(("nothing", "white"))
        for row in range(height + 5, sheet.max_row + 1):
            if sheet.cell(row, 1).value is None: break

            type_name = str(sheet.cell(row, 1).value)
            color_hex = sheet.cell(row, 1).fill.start_color.value
            color_hex = "#" + color_hex[2:]
            cells_states.append((type_name, color_hex))

        # read cell value and type
        for row in range(4, sheet.max_row):
            if sheet.cell(row, 1).value is None: break

            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row, col).value is None: break

                cell_value = float(sheet.cell(row, col).value)
                color_hex = sheet.cell(row, col).fill.start_color.value
                color_hex = "#" + color_hex[2:]
                cells[row - 4][col - 1] = (cell_value, color_hex)

        # Create cell objects
        cells_class: List[List[Cell]] = list(list())
        for y in range(height):
            cells_class.append(list())
            for x in range(width):
                cells_class[y].append(Cell(0, value=cells[y][x][0], example=example))
                for typee in range(len(cells_states)):
                    if cells_states[typee][1] == cells[y][x][1]:
                        cells_class[y][x].current_state = typee

        layer = Layer(name, cells_class, neighbourhood, cells_states)
        layers.append(layer)

        # Reading sliding window rules - optional
        rules = []
        rule_size = int((sheet.max_row - (height + 3 + len(cells_states))) / 4)
        for i in range(rule_size):
            match, result = [], []
            if neighbourhood == "von_neumann":
                match = read_von_neumann_excel(sheet, height + 7 + i * 4, 0, cells_states)
                result = read_von_neumann_excel(sheet, height + 7 + i * 4, 5, cells_states)
            elif neighbourhood == "moore":
                match = read_moore_excel(sheet, height + 7 + i * 4, 1, cells_states)
                result = read_moore_excel(sheet, height + 7 + i * 4, 5, cells_states)
            rule = (match, result)
            rules.append(rule)
        layer.rules = rules

    return layers


# Function that reads Board from either excel file or folder with text files
def read_board(folder: str, strategy: str) -> ModelBoard:
    layers = list()
    number_of_files = next(os.walk(folder))[2]
    for file in number_of_files:
        if file.endswith(".txt"):
            layer = read_layer_txt(os.path.join(folder, file))
            layers.append(layer)
        elif file.endswith(".xlsx"):
            layers = read_layer_excel(os.path.join(folder, file))
    return ModelBoard(layers, strategy)
